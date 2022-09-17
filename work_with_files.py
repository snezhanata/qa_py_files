import csv
import os
from zipfile import ZipFile, ZIP_DEFLATED
from PyPDF2 import PdfReader
from openpyxl import load_workbook

target_path = './resources/'
source_path = './source'
file_dir = os.listdir(source_path)


# Task_1: Запаковать в zip архив несколько разных файлов: pdf, xlsx, csv
# Task_2: Положить его в ресурсы
def file_zip_creation(file_zip_name):
    with ZipFile(f'{target_path}{file_zip_name}', 'w', compression=ZIP_DEFLATED) as file_zip:
        for file in file_dir:
            add_file = os.path.join(source_path, file)
            file_zip.write(add_file, os.path.relpath(add_file, source_path))
        file_zip.extractall(target_path)


# Task_3: Реализовать чтение и проверку содержимого каждого файла из архива
def file_csv_check(file_csv_name):
    print('CSV Check:')
    with open(f'{target_path}{file_csv_name}') as file_csv:
        table = list(csv.reader(file_csv))
        for index, value in enumerate(table):
            if index == 11:
                print(value[4])
        assert 'Mara' == table[2][1]
        # for row in enumerate(table):
        #     print(row) # распечатать весь файл


def file_xlsx_check(file_xlsx_name):
    print('\nXLSX Check:')
    workbook = load_workbook(f'{target_path}{file_xlsx_name}')
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)  # OR print(sheet['B3'].value)
    assert sheet[23][1].value == 'Что нужно сделать, чтобы функция возвратила значение?'


def file_pdf_check(file_pdf_name):
    print('\nPDF Check:')
    pdf_reader = PdfReader(f'{target_path}{file_pdf_name}')
    page = pdf_reader.pages[1]
    text = page.extract_text()
    print(text)
    assert 'typing this stuff' in text


file_zip_creation('file_4.zip')
file_csv_check('file_3.csv')
file_xlsx_check('file_2.xlsx')
file_pdf_check('file_1.pdf')
